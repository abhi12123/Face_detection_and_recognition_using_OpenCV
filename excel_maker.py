import os
from pprint import pprint
import urllib.request
import os
import openpyxl
import time

dirlist = os.listdir("images")
print(dirlist)
wb=openpyxl.load_workbook('att.xlsx')
sheet=wb['Sheet1']
start=sheet.max_row+2
c=start
sheet.cell(row=c,column=1).value=time.strftime("%d/%m/%Y")
sheet.cell(row=c,column=3).value=time.strftime("%H:%M:%S")
c=c+1

sheet.cell(row=c,column=1).value="Sno"
sheet.cell(row=c,column=2).value="Name"
sheet.cell(row=c,column=3).value="Present"
c=c+1
sl=1
for i in dirlist:
	sheet.cell(row=c,column=1).value=sl
	sheet.cell(row=c,column=2).value=i
	sheet.cell(row=c,column=3).value='no'
	c=c+1
	sl=sl+1

wb.save("att.xlsx")