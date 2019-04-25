#program to test openpyxl

import openpyxl
import time
wb=openpyxl.load_workbook('att.xlsx')
sheet=wb['Sheet1']
i=sheet.max_row+2
sheet.cell(row=i,column=1).value=time.strftime("%d/%m/%Y")
sheet.cell(row=i,column=3).value=time.strftime("%H:%M:%S")
wb.save("att.xlsx")