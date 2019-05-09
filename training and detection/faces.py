#program to detect face and register attedance

import numpy as np
import cv2
import pickle
import time
import urllib.request
import os
import openpyxl

face_cascade = cv2.CascadeClassifier('cascades/data/haarcascade_frontalface_alt2.xml')


recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("./recognizers/face-trainner.yml")

labels = {"person_name": 1}
with open("pickles/face-labels.pickle", 'rb') as f:
	og_labels = pickle.load(f)
	labels = {v:k for k,v in og_labels.items()}


wb=openpyxl.load_workbook('att.xlsx')
sheet=wb['Sheet1']
start=sheet.max_row+2
c=start
sheet.cell(row=c,column=1).value=time.strftime("%d/%m/%Y")
sheet.cell(row=c,column=3).value=time.strftime("%H:%M:%S")
c=c+1

sheet.cell(row=c,column=1).value="Sno"
sheet.cell(row=c,column=2).value="Name"
sheet.cell(row=c,column=3).value="Present"
c=c+1

for num,name in labels.items():
	sheet.cell(row=c,column=1).value=num
	sheet.cell(row=c,column=2).value=name
	sheet.cell(row=c,column=3).value='no'
	c=c+1

cap = cv2.VideoCapture(0)

ar=[0]*100

#url='http://192.168.1.2:8080/shot.jpg' #remove hash if ip webapp is being used
while(True):
	#imgResp=urllib.request.urlopen(url) #remove hash if ip webapp is being used
	#imgNp=np.array(bytearray(imgResp.read()),dtype=np.uint8) #remove hash if ip webapp is being used
	#frame=cv2.imdecode(imgNp,-1) #remove hash if ip webapp is being used
	ret, frame = cap.read() #remove hash if laptop camera is being used
	gray  = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
	faces = face_cascade.detectMultiScale(gray, scaleFactor=1.5, minNeighbors=5)
	for (x, y, w, h) in faces:
		roi_gray = gray[y:y+h, x:x+w] 
		roi_color = frame[y:y+h, x:x+w]
		id_, conf = recognizer.predict(roi_gray)
		font = cv2.FONT_HERSHEY_SIMPLEX
		color = (255, 255, 255)
		stroke = 2
		if conf>=40 and conf <= 85:
			ar[id_]=ar[id_]+1
			name = labels[id_]
			cv2.putText(frame, name, (x,y), font, 1, color, stroke, cv2.LINE_AA)
			if(ar[id_]==30):
				print (name," is present")
				sheet.cell(row=start+id_+1,column=3).value='yes'
		else:
			cv2.putText(frame, "unknown", (x,y), font, 1, color, stroke, cv2.LINE_AA)
		color = (255, 0, 0)
		stroke = 2
		end_cord_x = x + w
		end_cord_y = y + h
		cv2.rectangle(frame, (x, y), (end_cord_x, end_cord_y), color, stroke)
	cv2.imshow('frame',frame)
	if cv2.waitKey(20) & 0xFF == ord('q'):
		break
wb.save("att.xlsx")
cap.release()
cv2.destroyAllWindows()
